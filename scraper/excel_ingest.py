"""Ingest official e-auksion Excel report into AuksionWatch DB.

Source: D:/hackaton/23131.xlsx — Farg'ona viloyati 9266 lots
Schema (row 4 headers):
  A Manzil
  B Bino maydoni kv.m.
  C Baholangan narx
  D Ijara maydoni
  E Oxirgi auksion sanasi
  F Lot raqami
  G Viloyat
  H Tuman/Shahar
  I Kategoriya nomi
  J T/R
  K Boshlang'ich narx
  L Buyurtmachi nomi (seller)
  M Sozlamalar
  N Mulk turi
  O Yer maydoni (ga)
  P Necha marta savdoga chiqarilganligi (re-auction count)
"""
import sys
import json
import re
import hashlib
from collections import Counter, defaultdict
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import load_workbook
from sqlmodel import Session, select, func
from backend.db import engine, init_db
from backend.models import Lot

ROOT = Path(__file__).parent.parent
EXCEL_PATH = Path("D:/hackaton/23131.xlsx")

REGION_CENTROIDS = {
    "UZ-FA": (40.3834, 71.7842),  # Farg'ona
    "UZ-TK": (41.3111, 69.2797),
    "UZ-AN": (40.7821, 72.3442),
    "UZ-BU": (39.7747, 64.4286),
}

REGION_NAME_TO_CODE = {
    "Farg`ona viloyati": "UZ-FA",
    "Farg'ona viloyati": "UZ-FA",
    "Toshkent shahri": "UZ-TK",
    "Toshkent viloyati": "UZ-TO",
    "Andijon viloyati": "UZ-AN",
    "Buxoro viloyati": "UZ-BU",
}


def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(" ", "").replace(",", ".")
    s = re.sub(r"[^\d.]", "", s)
    try:
        return float(s) if s else None
    except ValueError:
        return None


def seller_hint_from_name(name: str | None) -> str:
    if not name:
        return "unknown"
    # Farg'ona excel — all are individual buyurtmachi (seller). Categorize:
    n = name.upper()
    if "DAVAKTIV" in n or "ДАВАКТИВ" in n:
        return "davaktiv"
    if "BANK" in n or "БАНК" in n:
        return "bank"
    return "individual"


def seller_id_from_name(name: str | None) -> int | None:
    """Stable hash → integer seller ID for graph analysis."""
    if not name:
        return None
    h = hashlib.md5(name.encode("utf-8")).hexdigest()
    return int(h[:8], 16) % (10**9)


def jitter_geo(region_code, lot_id):
    if region_code not in REGION_CENTROIDS:
        return None, None
    lat, lon = REGION_CENTROIDS[region_code]
    seed = (lot_id or 0) % 1000
    return (
        lat + ((seed % 200) - 100) / 250,
        lon + (((seed * 7) % 200) - 100) / 250,
    )


def evaluate_excel_risk(row: dict, seller_stats: dict, region_stats: dict) -> dict:
    """Risk engine for Excel rows. New signals from Farg'ona dataset."""
    flags = []

    # 1. Re-auction count — most powerful signal in this dataset
    times = row.get("times_auctioned") or 0
    if times >= 15:
        flags.append({
            "type": "many_reauctions",
            "score": 35,
            "title": f"{int(times)} marta auksionga chiqarilgan",
            "desc": "Lot ko'p marta sotilmagan — narx sun'iy yuqori yoki sotuvchi g'olibni kutmoqda.",
        })
    elif times >= 8:
        flags.append({
            "type": "repeat_auction",
            "score": 22,
            "title": f"{int(times)} marta auksionga chiqarilgan",
            "desc": "Lot bir necha marta sotilmagan — shubhali takroriy auksion.",
        })
    elif times >= 5:
        flags.append({
            "type": "reauction",
            "score": 12,
            "title": f"{int(times)} marta auksionga chiqarilgan",
            "desc": "Bir necha marta sotilmagan lot.",
        })

    # 2. Seller concentration — if one seller has > 1000 lots in single region
    seller_total = seller_stats.get(row.get("seller_id"), {}).get("total", 0)
    if seller_total >= 1000:
        flags.append({
            "type": "monopoly_seller",
            "score": 30,
            "title": "Monopoliyalashgan sotuvchi",
            "desc": f"Bu sotuvchining {seller_total} ta lot — hududda hukmron mavqe.",
        })
    elif seller_total >= 300:
        flags.append({
            "type": "dominant_seller",
            "score": 18,
            "title": "Hukmron sotuvchi",
            "desc": f"Bu sotuvchining {seller_total} ta lot — sezilarli mavqe.",
        })

    # 3. Underpriced vs region median (use start_price)
    sp = row.get("start_price")
    region = row.get("region")
    if sp and region in region_stats:
        rs = region_stats[region]
        if rs.get("median") and sp < rs["median"] * 0.3:
            flags.append({
                "type": "deeply_underpriced",
                "score": 22,
                "title": "Hudud medianidan keskin past narx",
                "desc": f"Boshlang'ich narx hudud medianidan {(1-sp/rs['median'])*100:.0f}% past.",
            })

    # 4. Appraised price vs start price gap
    appraised = row.get("appraised_price")
    if appraised and sp and appraised > 0:
        ratio = sp / appraised
        if ratio < 0.5:
            flags.append({
                "type": "below_appraisal",
                "score": 18,
                "title": "Baholangan narxdan past boshlanish",
                "desc": f"Boshlang'ich narx baholangandan {(1-ratio)*100:.0f}% past.",
            })

    # 5. Many re-auctions + dominant seller (combo signal)
    if times >= 8 and seller_total >= 300:
        flags.append({
            "type": "stuck_lot_pattern",
            "score": 10,
            "title": "Yopishgan lot + hukmron sotuvchi",
            "desc": "Takroriy auksion va hukmron sotuvchi birgalikda — kelishuv ehtimoli yuqori.",
        })

    score = min(100, sum(f["score"] for f in flags))
    level = "low" if score < 40 else ("medium" if score < 70 else "high")
    return {"score": score, "level": level, "flags": flags}


def main():
    init_db()
    print(f"[excel] reading {EXCEL_PATH}")
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=4, values_only=True))
    print(f"[excel] {len(rows)} rows")

    # Pass 1: build cleaned records + statistics
    records = []
    seller_counts = defaultdict(lambda: {"total": 0, "names": set()})
    region_prices = defaultdict(list)

    for r in rows:
        if not r or not r[5]:  # lot raqami
            continue
        lot_id_raw = r[5]
        try:
            lot_id = int(lot_id_raw)
        except (TypeError, ValueError):
            continue
        if lot_id < 1000000:
            continue  # skip headers/junk

        address = r[0]
        building_area = num(r[1])
        appraised = num(r[2])
        rental_area = r[3]
        last_auction = r[4]
        region_name = r[6]
        district = r[7]
        category = r[8]
        start_price = num(r[10])
        seller_name = r[11]
        settings = r[12]
        property_type = r[13]
        land_area = r[14]
        times_auctioned = num(r[15])

        region_code = REGION_NAME_TO_CODE.get(region_name, "UZ-FA")
        seller_id = seller_id_from_name(seller_name)

        rec = {
            "lot_id": lot_id,
            "title": f"{property_type or category or 'Lot'} — {district or ''}".strip(" —"),
            "lot_type": property_type,
            "lot_type_specific": category,
            "address": address,
            "region": region_code,
            "district": district,
            "start_price": start_price,
            "appraised_price": appraised,
            # Excel hisobotida yopiq/ochiq ustuni yo'q. Faqat API'dan tasdiqlanadi.
            "auction_type": "unknown",
            "seller_name": seller_name,
            "seller_id": seller_id,
            "seller_hint": seller_hint_from_name(seller_name),
            "times_auctioned": int(times_auctioned) if times_auctioned else None,
            "end_time": str(last_auction) if last_auction else None,
            "status": "savdoda" if times_auctioned and times_auctioned >= 1 else None,
            "settings": settings,
            "building_area": building_area,
            "rental_area": str(rental_area) if rental_area else None,
            "land_area": str(land_area) if land_area else None,
        }
        records.append(rec)

        if seller_id:
            seller_counts[seller_id]["total"] += 1
            seller_counts[seller_id]["names"].add(seller_name)
        if start_price and region_code:
            region_prices[region_code].append(start_price)

    # region stats
    import statistics
    region_stats = {
        r: {"median": statistics.median(prices), "n": len(prices)}
        for r, prices in region_prices.items()
        if len(prices) >= 5
    }
    print(f"[excel] sellers: {len(seller_counts)}, region_stats: {len(region_stats)}")

    # Top sellers report
    top_sellers = sorted(
        seller_counts.items(),
        key=lambda x: -x[1]["total"],
    )[:10]
    print("[excel] TOP 10 sellers by lot count:")
    for sid, info in top_sellers:
        names = list(info["names"])[0] if info["names"] else "?"
        print(f"  {info['total']:5}  {names[:50]}")

    # Pass 2: write to DB
    inserted = 0
    skipped = 0
    flag_levels = Counter()

    with Session(engine) as session:
        for rec in records:
            existing = session.get(Lot, rec["lot_id"])
            if existing:
                skipped += 1
                continue

            risk = evaluate_excel_risk(rec, seller_counts, region_stats)
            flag_levels[risk["level"]] += 1

            lat, lon = jitter_geo(rec["region"], rec["lot_id"])

            ai_summary = None
            if risk["level"] == "high":
                titles = ", ".join(f["title"] for f in risk["flags"])
                ai_summary = f"YUQORI XAVF: {titles}. Takroriy auksion + hukmron sotuvchi sxemasiga o'xshash."
            elif risk["level"] == "medium":
                ai_summary = f"O'RTA XAVF: {len(risk['flags'])} ta shubhali belgi mavjud."

            lot = Lot(
                id=rec["lot_id"],
                url=f"https://e-auksion.uz/lot-view?lot_id={rec['lot_id']}",
                title=rec["title"],
                lot_type=rec["lot_type"],
                lot_type_specific=rec["lot_type_specific"],
                address=rec["address"],
                region=rec["region"],
                district=rec["district"],
                start_price=rec["start_price"],
                sold_price=None,
                deposit=None,
                step_price=None,
                installment_months=None,
                auction_method="Auksion",
                auction_style=None,
                auction_type=rec["auction_type"],
                start_time=None,
                deadline=None,
                end_time=rec["end_time"],
                status=rec["status"],
                views=None,
                bidders_count=None,
                seller_hint=rec["seller_hint"],
                geo_lat=lat,
                geo_lon=lon,
                risk_score=risk["score"],
                risk_level=risk["level"],
                ai_summary=ai_summary,
                flags=risk["flags"],
            )
            session.add(lot)
            inserted += 1

            if inserted % 500 == 0:
                session.commit()
                print(f"  inserted {inserted}...")

        session.commit()

    print(f"\n[excel] DONE")
    print(f"  inserted: {inserted}")
    print(f"  skipped (already in DB): {skipped}")
    print(f"  risk levels: {dict(flag_levels)}")

    # final DB stats
    with Session(engine) as session:
        total = session.exec(select(func.count(Lot.id))).one()
        high = session.exec(select(func.count(Lot.id)).where(Lot.risk_level == "high")).one()
        medium = session.exec(select(func.count(Lot.id)).where(Lot.risk_level == "medium")).one()
        print(f"\n[db] total: {total}, high: {high}, medium: {medium}")


if __name__ == "__main__":
    main()
