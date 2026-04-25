"""Backfill Excel-imported lots with new fields (appraised_price, times_auctioned, seller_id).

The first excel_ingest didn't populate these because the model didn't have them yet.
Now we re-read Excel and UPDATE existing rows in DB.
"""
import sys
import hashlib
import re
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

from openpyxl import load_workbook
from sqlmodel import Session, select
from backend.db import engine
from backend.models import Lot

EXCEL_PATH = Path("D:/hackaton/23131.xlsx")


def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(" ", " ").replace(",", ".")
    s = re.sub(r"[^\d.]", "", s)
    try:
        return float(s) if s else None
    except ValueError:
        return None


def seller_id_from_name(name):
    if not name:
        return None
    h = hashlib.md5(name.encode("utf-8")).hexdigest()
    return int(h[:8], 16) % (10**9)


def main():
    print(f"[backfill] reading {EXCEL_PATH}")
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    excel_data = {}
    for r in ws.iter_rows(min_row=4, values_only=True):
        if not r or not r[5]:
            continue
        try:
            lot_id = int(r[5])
        except (TypeError, ValueError):
            continue
        if lot_id < 1000000:
            continue
        excel_data[lot_id] = {
            "appraised_price": num(r[2]),
            "seller_name": r[11],
            "seller_id": seller_id_from_name(r[11]),
            "times_auctioned": int(num(r[15])) if num(r[15]) else None,
        }

    print(f"[backfill] excel rows: {len(excel_data)}")

    updated = 0
    with Session(engine) as session:
        all_lots = session.exec(select(Lot)).all()
        print(f"[backfill] DB rows: {len(all_lots)}")
        for lot in all_lots:
            if lot.id not in excel_data:
                continue
            d = excel_data[lot.id]
            if lot.appraised_price is None and d["appraised_price"]:
                lot.appraised_price = d["appraised_price"]
            if lot.times_auctioned is None and d["times_auctioned"]:
                lot.times_auctioned = d["times_auctioned"]
            if lot.seller_id is None and d["seller_id"]:
                lot.seller_id = d["seller_id"]
            if lot.seller_name is None and d["seller_name"]:
                lot.seller_name = d["seller_name"]
            updated += 1
            if updated % 1000 == 0:
                session.commit()
                print(f"  updated {updated}...")
        session.commit()

    print(f"[backfill] DONE: {updated} lots updated with Excel fields")


if __name__ == "__main__":
    main()
